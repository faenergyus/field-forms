"""
FAE XSPOC History API Server
Serves well trending data from tblDataHistory via FastAPI.
Runs on svrxspoc at C:/AI/ANALYST/data/api_server.py

Endpoints:
  GET /health              - Health check
  GET /wells               - List enabled wells with POCType
  GET /history/{well}      - Trending data (params: days, addresses)
  GET /registers/{well}    - Available registers for a well
"""

import json
import logging
import logging.handlers
import struct
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import pyodbc
from fastapi import FastAPI, Query, HTTPException, Depends, Header, UploadFile, File, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import ORJSONResponse, Response
from google.oauth2 import id_token
from google.auth.transport import requests as google_requests

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
CONN_STR = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=sql2;"
    "DATABASE=xspoc;"
    "UID=fae_api_read;"
    "PWD=FAE_xspoc_r3ad!2026;"
)
# Ops_Reporting connection — wbd_* tables + Pumper_Data_Calcs
OPS_CONN_STR = (
    r"DRIVER={ODBC Driver 17 for SQL Server};"
    r"SERVER=sf\sqldev;"
    r"DATABASE=Ops_Reporting;"
    r"Trusted_Connection=yes;"
    r"TrustServerCertificate=yes;"
)
LOG_DIR = Path(r"C:\AI\ANALYST\data\logs")
LOG_DIR.mkdir(parents=True, exist_ok=True)
PORT = 8080
GOOGLE_CLIENT_ID = "629380732564-eh0uasomqjdo3up8o2po685g9ep544iq.apps.googleusercontent.com"

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
log = logging.getLogger("xspoc_api")
log.setLevel(logging.INFO)
_fmt = logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S")

_fh = logging.handlers.RotatingFileHandler(
    LOG_DIR / "api_server.log", maxBytes=2_000_000, backupCount=3, encoding="utf-8"
)
_fh.setFormatter(_fmt)
log.addHandler(_fh)

if sys.stdout is not None:
    _ch = logging.StreamHandler()
    _ch.setFormatter(_fmt)
    log.addHandler(_ch)

# ---------------------------------------------------------------------------
# Register mapping: (POCType, Address) -> (param_name, scale, unit)
# ---------------------------------------------------------------------------
REGISTER_MAP = {
    # Lufkin SAM (POCType=8)
    (8, 32606): ("InferredProd", 0.1, "bbl/d"),
    (8, 32611): ("PctRun", 1.0, "%"),
    (8, 32614): ("Fillage", 0.01, "%"),
    (8, 32569): ("PPRL", 1.0, "lbs"),
    (8, 32572): ("MPRL", 1.0, "lbs"),
    (8, 32613): ("FluidLoad", 1.0, "lbs"),
    (8, 3002):  ("SPM", 0.1, "spm"),
    # ChampionX SMARTEN (POCType=17)
    (17, 30055): ("InferredProd", 1.0, "bbl/d"),
    (17, 30067): ("PctRun", 1.0, "%"),
    (17, 30061): ("Fillage", 1.0, "%"),
    (17, 30077): ("PPRL", 1.0, "lbs"),
    (17, 30079): ("MPRL", 1.0, "lbs"),
    (17, 30166): ("FluidLoad", 1.0, "lbs"),
    (17, 30168): ("SPM", 1.0, "spm"),
    # Weatherford Well Pilot (POCType=16)
    (16, 811):  ("InferredProd", 1.0, "bbl/d"),
    (16, 1276): ("PctRun", 1.0, "%"),
    (16, 3149): ("Fillage", 1.0, "%"),
    (16, 80):   ("PPRL", 1.0, "lbs"),
    (16, 87):   ("MPRL", 1.0, "lbs"),
    (16, 3147): ("FluidLoad", 1.0, "lbs"),
}

# Reverse lookup: param_name -> {poc_type: address}
PARAM_ADDRESSES = {}
for (poc, addr), (pname, _, _) in REGISTER_MAP.items():
    PARAM_ADDRESSES.setdefault(pname, {})[poc] = addr


def get_conn():
    return pyodbc.connect(CONN_STR, timeout=30)


def get_ops_conn():
    return pyodbc.connect(OPS_CONN_STR, timeout=30)


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
_google_request = google_requests.Request()


async def get_current_user(authorization: Optional[str] = Header(None)) -> str:
    """Verify Google ID token from Authorization: Bearer <token> header."""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing or invalid Authorization header")
    token = authorization[7:]
    try:
        idinfo = id_token.verify_oauth2_token(token, _google_request, GOOGLE_CLIENT_ID)
        email = idinfo.get("email", "")
        if not email:
            raise ValueError("No email in token")
        log.info("Auth OK: %s", email)
        return email
    except ValueError as e:
        raise HTTPException(status_code=401, detail=f"Invalid token: {e}")


# ---------------------------------------------------------------------------
# FastAPI app
# ---------------------------------------------------------------------------
app = FastAPI(title="FAE XSPOC API", version="1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://faenergyus.github.io", "http://localhost:8000", "http://127.0.0.1:8000"],
    allow_origin_regex=r"https://.*\.trycloudflare\.com",
    allow_methods=["GET", "PUT", "POST", "DELETE"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# KML — only external_wells.kml (third-party wells the user maintains by hand).
# FAE wells flow through xlsx → SQL → JSON → map; no need to also represent
# them as KML.
# ---------------------------------------------------------------------------
KML_DIR = Path(r"C:\AI\ANALYST\data\kml")

EXTERNAL_KML_STUB = (
    '<?xml version="1.0" encoding="UTF-8"?>\n'
    '<kml xmlns="http://www.opengis.net/kml/2.2"><Document>'
    '<name>External Wells (third party)</name>'
    '<description>Manually maintained — wells not in Pumper_Data_Calcs '
    '(offset operators, scout wells, future locations, etc.). Edit in '
    'Google Earth Pro, upload back via PUT /kml/external_wells.kml.</description>'
    '<Style id="external_style"><IconStyle>'
    '<color>ff888888</color><scale>0.8</scale>'
    '<Icon><href>http://maps.google.com/mapfiles/kml/shapes/open-diamond.png</href></Icon>'
    '</IconStyle><LabelStyle><scale>0.7</scale></LabelStyle></Style>'
    '</Document></kml>'
)


def _ensure_external_kml():
    """Create external_wells.kml stub on first boot. Never overwrite."""
    KML_DIR.mkdir(parents=True, exist_ok=True)
    ext = KML_DIR / "external_wells.kml"
    if not ext.exists():
        ext.write_text(EXTERNAL_KML_STUB, encoding="utf-8")
        log.info("Seeded external_wells.kml stub")


_ensure_external_kml()


def _safe_kml_path(name: str) -> Path:
    if not name.endswith(".kml") or "/" in name or "\\" in name or ".." in name:
        raise HTTPException(400, "bad filename")
    return KML_DIR / name


@app.get("/kml/{name}")
def get_kml(name: str):
    """Download a KML file from the kml directory (e.g. external_wells.kml)."""
    path = _safe_kml_path(name)
    if not path.exists():
        raise HTTPException(404, "not found")
    return Response(content=path.read_bytes(),
                    media_type="application/vnd.google-earth.kml+xml",
                    headers={"Content-Disposition": f'inline; filename="{name}"'})


@app.put("/kml/{name}")
async def put_kml(name: str, request: Request, user: str = Depends(get_current_user)):
    """Upload an edited KML file. Auth-gated. Body is the raw KML bytes."""
    path = _safe_kml_path(name)
    body = await request.body()
    if not body or not body.lstrip().startswith(b"<"):
        raise HTTPException(400, "body doesn't look like KML")
    # quick well-formedness check
    try:
        from xml.etree import ElementTree as ET
        ET.fromstring(body)
    except Exception as e:
        raise HTTPException(400, f"invalid XML: {e}")
    path.write_bytes(body)
    log.info("KML uploaded: %s (%d bytes) by %s", name, len(body), user)
    return {"ok": True, "file": name, "bytes": len(body)}


@app.get("/health")
def health():
    """Health check — also verifies DB connectivity."""
    try:
        conn = get_conn()
        conn.cursor().execute("SELECT 1")
        conn.close()
        return {"status": "ok", "db": "connected", "time": datetime.now().isoformat()}
    except Exception as e:
        return {"status": "degraded", "db": str(e), "time": datetime.now().isoformat()}


def decode_card_binary(blob):
    """Decode an XSPOC card binary field (N single-precision floats, first
    half = load, second half = position). Returns {position, load} or None."""
    import struct as _struct
    if blob is None: return None
    try:
        b = bytes(blob)
    except Exception:
        return None
    if len(b) < 8 or len(b) % 4 != 0: return None
    n = len(b) // 4
    floats = _struct.unpack(f"<{n}f", b)
    half = n // 2
    return {"load": list(floats[:half]), "position": list(floats[half:])}


@app.get("/scada/dyno-trend/{node_id:path}")
def get_dyno_trend(
    node_id: str,
    days: int = Query(default=14, ge=1, le=90),
    user: str = Depends(get_current_user),
):
    """Per-card max/min surface load + runtime + fillage for a well, last N days.

    Reads tblCardData directly so the SCADA-tab trend chart doesn't depend on
    cards.json (which only carries a small recent slice). Surface card binary
    is decoded server-side to compute peak/min polished-rod load.
    """
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT [Date], NodeID, SurfaceCardB, Runtime, Fillage
        FROM tblCardData
        WHERE NodeID = ?
          AND [Date] >= DATEADD(day, -?, GETDATE())
          AND SurfaceCardB IS NOT NULL
        ORDER BY [Date] ASC
    """, node_id, days)

    out = []
    for dt, nid, surf_b, runtime, fillage in cur.fetchall():
        # decode_card_binary returns {position, load} or None
        card = None
        try:
            card = decode_card_binary(surf_b)
        except Exception:
            card = None
        if not card or not card.get("load"):
            continue
        loads = card["load"]
        out.append({
            "date":      dt.isoformat() if dt else None,
            "max_load":  max(loads),
            "min_load":  min(loads),
            "runtime":   float(runtime) if runtime is not None else None,
            "fillage":   float(fillage) if fillage is not None else None,
        })
    conn.close()
    return {"node_id": node_id, "days": days, "rows": out}


@app.get("/wells")
def list_wells(user: str = Depends(get_current_user)):
    """List all enabled wells with POCType."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT NodeID, POCType, Enabled, RunStatus, CommStatus,
               PumpFillage, InferredProd, LastGoodScanTime
        FROM tblNodeMaster
        WHERE Enabled = 1 AND POCType IN (8, 16, 17)
        ORDER BY NodeID
    """)
    wells = []
    for row in cur.fetchall():
        wells.append({
            "well": row[0], "pocType": row[1], "enabled": row[2],
            "runStatus": row[3], "commStatus": row[4],
            "fillage": row[5], "inferredProd": row[6],
            "lastScan": row[7].isoformat() if row[7] else None,
        })
    conn.close()
    return {"wells": wells, "count": len(wells)}


@app.get("/history/{well}", response_class=ORJSONResponse)
def get_history(
    well: str,
    days: int = Query(default=7, ge=1, le=365, description="Days of history"),
    params: Optional[str] = Query(
        default=None,
        description="Comma-separated param names (e.g. Fillage,PPRL,MPRL). Default: all."
    ),
    max_points: Optional[int] = Query(
        default=None, ge=100, le=50000,
        description="Downsample to this many points per param. Default: no limit."
    ),
    user: str = Depends(get_current_user),
):
    """
    Get trending history for a well.

    Returns {x: epoch_ms, y: value} arrays per parameter, ready for Chart.js.
    """
    conn = get_conn()
    cur = conn.cursor()

    # Get POCType for this well
    cur.execute("SELECT POCType FROM tblNodeMaster WHERE NodeID = ?", well)
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, f"Well '{well}' not found")
    poc_type = row[0]

    # Build address map for this POC type
    addr_map = {}
    for (pt, addr), (pname, scale, unit) in REGISTER_MAP.items():
        if pt == poc_type:
            addr_map[addr] = (pname, scale, unit)

    # Filter to requested params if specified
    if params:
        requested = set(p.strip() for p in params.split(","))
        addr_map = {a: v for a, v in addr_map.items() if v[0] in requested}

    if not addr_map:
        conn.close()
        raise HTTPException(400, f"No valid parameters for POCType {poc_type}")

    addr_list = sorted(addr_map.keys())
    placeholders = ",".join(str(a) for a in addr_list)

    cur.execute(f"""
        SELECT [Date], Address, Value
        FROM tblDataHistory WITH (NOLOCK)
        WHERE NodeID = ?
          AND Address IN ({placeholders})
          AND [Date] > DATEADD(day, -{days}, GETDATE())
        ORDER BY [Date]
    """, well)

    # Collect into per-param arrays
    param_data = {}
    for row in cur.fetchall():
        dt, addr, val = row
        if addr not in addr_map:
            continue
        pname, scale, unit = addr_map[addr]
        epoch_ms = int(dt.timestamp() * 1000)
        scaled = round(val * scale, 2) if scale != 1.0 else round(val, 2)
        entry = param_data.setdefault(pname, {"unit": unit, "data": []})
        entry["data"].append([epoch_ms, scaled])

    conn.close()

    # Downsample if requested (simple nth-point for speed)
    if max_points:
        for pname, pinfo in param_data.items():
            data = pinfo["data"]
            if len(data) > max_points:
                step = len(data) / max_points
                pinfo["data"] = [data[int(i * step)] for i in range(max_points)]

    return {
        "well": well,
        "pocType": poc_type,
        "days": days,
        "params": param_data,
    }


@app.get("/registers/{well}")
def get_registers(well: str, user: str = Depends(get_current_user)):
    """List all addresses currently being logged for a well (last 24h)."""
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT POCType FROM tblNodeMaster WHERE NodeID = ?", well)
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, f"Well '{well}' not found")
    poc_type = row[0]

    cur.execute("""
        SELECT Address, COUNT(*) as cnt,
               MIN(Value) as minv, MAX(Value) as maxv, AVG(Value) as avgv
        FROM tblDataHistory WITH (NOLOCK)
        WHERE NodeID = ?
          AND [Date] > DATEADD(day, -1, GETDATE())
        GROUP BY Address
        HAVING COUNT(*) > 10
        ORDER BY Address
    """, well)

    registers = []
    for row in cur.fetchall():
        addr = row[0]
        mapped = REGISTER_MAP.get((poc_type, addr))
        registers.append({
            "address": addr,
            "name": mapped[0] if mapped else None,
            "unit": mapped[2] if mapped else None,
            "count_24h": row[1],
            "min": round(row[2], 2),
            "max": round(row[3], 2),
            "avg": round(row[4], 2),
        })
    conn.close()
    return {"well": well, "pocType": poc_type, "registers": registers}


# ---------------------------------------------------------------------------
# Dyno card history (full list metadata + per-card binary on demand)
# ---------------------------------------------------------------------------
def _decode_card_blob(blob):
    """N single-precision floats; first half = load, second half = position."""
    if blob is None or len(blob) < 8 or len(blob) % 4 != 0:
        return None
    n = len(blob) // 4
    floats = struct.unpack(f"<{n}f", blob)
    half = n // 2
    return {"position": list(floats[half:]), "load": list(floats[:half])}


@app.get("/card-list/{well}", response_class=ORJSONResponse)
def card_list(
    well: str,
    limit: int = Query(default=500, ge=1, le=10000),
    days: Optional[int] = Query(default=None, ge=1, le=3650),
    user: str = Depends(get_current_user),
):
    """List card metadata (no binary) for a well, most recent first."""
    conn = get_conn()
    cur = conn.cursor()
    where = "NodeID = ?"
    params = [well]
    if days is not None:
        where += f" AND [Date] > DATEADD(day, -{int(days)}, GETDATE())"
    cur.execute(
        f"""
        SELECT TOP {int(limit)} [Date], CardType, SPM, StrokeLength, Fillage,
               Area, CardArea, Runtime, CauseID, AnalysisDate
        FROM tblCardData WITH (NOLOCK)
        WHERE {where}
        ORDER BY [Date] DESC
        """,
        *params,
    )
    rows = []
    for r in cur.fetchall():
        rows.append({
            "Date": r[0].isoformat() if r[0] else None,
            "CardType": r[1],
            "SPM": float(r[2]) if r[2] is not None else None,
            "StrokeLength": float(r[3]) if r[3] is not None else None,
            "Fillage": float(r[4]) if r[4] is not None else None,
            "Area": float(r[5]) if r[5] is not None else None,
            "CardArea": float(r[6]) if r[6] is not None else None,
            "Runtime": float(r[7]) if r[7] is not None else None,
            "CauseID": r[8],
            "AnalysisDate": r[9].isoformat() if r[9] else None,
        })
    conn.close()
    return {"well": well, "count": len(rows), "cards": rows}


@app.get("/card/{well}/{card_date}", response_class=ORJSONResponse)
def card_detail(
    well: str,
    card_date: str,
    user: str = Depends(get_current_user),
):
    """Fetch one card's binaries (Surface, Downhole, POC DH) by exact Date."""
    try:
        dt = datetime.fromisoformat(card_date.replace("Z", ""))
    except ValueError:
        raise HTTPException(400, "Invalid card_date (ISO 8601 expected)")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT TOP 1 [Date], CardType, SPM, StrokeLength, Fillage, Area, CardArea,
               Runtime, CauseID, AnalysisDate,
               SurfaceCardB, DownholeCardB, POCDownholeCardB,
               PredictedCardB, PermissibleLoadUpB, PermissibleLoadDownB
        FROM tblCardData WITH (NOLOCK)
        WHERE NodeID = ? AND [Date] = ?
        """,
        well, dt,
    )
    r = cur.fetchone()
    conn.close()
    if not r:
        raise HTTPException(404, "Card not found")
    return {
        "well": well,
        "Date": r[0].isoformat() if r[0] else None,
        "CardType": r[1],
        "SPM": float(r[2]) if r[2] is not None else None,
        "StrokeLength": float(r[3]) if r[3] is not None else None,
        "Fillage": float(r[4]) if r[4] is not None else None,
        "Area": float(r[5]) if r[5] is not None else None,
        "CardArea": float(r[6]) if r[6] is not None else None,
        "Runtime": float(r[7]) if r[7] is not None else None,
        "CauseID": r[8],
        "AnalysisDate": r[9].isoformat() if r[9] else None,
        "SurfaceCardB": _decode_card_blob(r[10]),
        "DownholeCardB": _decode_card_blob(r[11]),
        "POCDownholeCardB": _decode_card_blob(r[12]),
        "PredictedCardB": _decode_card_blob(r[13]),
        "PermissibleLoadUpB": _decode_card_blob(r[14]),
        "PermissibleLoadDownB": _decode_card_blob(r[15]),
    }


# ---------------------------------------------------------------------------
# Chart config persistence (per user + per well)
# ---------------------------------------------------------------------------
import hashlib

CHART_CONFIG_DIR = Path(r"C:\AI\ANALYST\data\chart_configs")
CHART_CONFIG_DIR.mkdir(parents=True, exist_ok=True)


def _config_path(email: str, well: str) -> Path:
    safe_email = hashlib.md5(email.lower().encode()).hexdigest()[:12]
    safe_well = well.replace(" ", "_").replace("#", "")
    user_dir = CHART_CONFIG_DIR / safe_email
    user_dir.mkdir(exist_ok=True)
    return user_dir / f"{safe_well}.json"


@app.get("/chart-config/{well}")
def get_chart_config(well: str, user: str = Depends(get_current_user)):
    """Load saved chart layout: per-well first, then user default, then nothing."""
    # Per-well config
    path = _config_path(user, well)
    if path.exists():
        data = json.loads(path.read_text(encoding="utf-8"))
        data["source"] = "well"
        return data
    # User default template
    default_path = _config_path(user, "_default")
    if default_path.exists():
        data = json.loads(default_path.read_text(encoding="utf-8"))
        data["source"] = "default"
        return data
    return {"panels": None, "source": "none"}


from pydantic import BaseModel


class ChartConfigBody(BaseModel):
    panels: list


@app.put("/chart-config/{well}")
def save_chart_config_put(well: str, body: ChartConfigBody, user: str = Depends(get_current_user)):
    """Save chart layout for this user + well."""
    path = _config_path(user, well)
    path.write_text(json.dumps({"panels": body.panels}, indent=2), encoding="utf-8")
    log.info("Chart config saved: %s for %s (%d panels)", well, user, len(body.panels))
    return {"saved": True, "panels": len(body.panels)}


@app.put("/chart-config-default")
def save_default_config(body: ChartConfigBody, user: str = Depends(get_current_user)):
    """Save default chart template for this user (applies to all wells without a custom config)."""
    path = _config_path(user, "_default")
    path.write_text(json.dumps({"panels": body.panels}, indent=2), encoding="utf-8")
    log.info("Default chart config saved for %s (%d panels)", user, len(body.panels))
    return {"saved": True, "panels": len(body.panels)}


# Arbitrary per-user UI prefs (cross-device) — diagnosis collapsed state, etc.
class UserPrefsBody(BaseModel):
    prefs: dict


def _prefs_path(email: str) -> Path:
    safe_email = hashlib.md5(email.lower().encode()).hexdigest()[:12]
    user_dir = CHART_CONFIG_DIR / safe_email
    user_dir.mkdir(exist_ok=True)
    return user_dir / "_prefs.json"


@app.get("/user-prefs")
def get_user_prefs(user: str = Depends(get_current_user)):
    """Get arbitrary UI preferences blob for the user."""
    path = _prefs_path(user)
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {}


@app.put("/user-prefs")
def save_user_prefs(body: UserPrefsBody, user: str = Depends(get_current_user)):
    """Replace the user's preferences blob."""
    path = _prefs_path(user)
    path.write_text(json.dumps(body.prefs, indent=2), encoding="utf-8")
    return {"saved": True}


# ---------------------------------------------------------------------------
# PBI Production (Analysis trends)
# ---------------------------------------------------------------------------
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

_pbi_token = None
_pbi_token_expires = 0


def _get_pbi_token():
    global _pbi_token, _pbi_token_expires
    import time as _time
    if _pbi_token and _time.time() < _pbi_token_expires - 60:
        return _pbi_token
    try:
        from pbi_helpers import get_delegated_token
        _pbi_token = get_delegated_token()
        _pbi_token_expires = _time.time() + 3500
        return _pbi_token
    except Exception as e:
        log.error("PBI token error: %s", e)
        raise HTTPException(503, f"PBI auth failed: {e}")


PROD_COLUMNS = {
    "Oil":      ("DP_OIL", "bbl"),
    "Gas":      ("DP_GAS", "MCF"),
    "Water":    ("DP_WATER", "bbl"),
    "InjWater": ("DP_INJWATER", "bbl"),
    "FTP":      ("DP_FTP", "psi"),
    "CP":       ("DP_CP", "psi"),
}


@app.get("/production/{well}", response_class=ORJSONResponse)
def get_production(
    well: str,
    days: int = Query(default=365, ge=1, le=3650, description="Days of production history"),
    params: Optional[str] = Query(
        default=None,
        description="Comma-separated param names (e.g. Oil,Gas,Water). Default: all."
    ),
    user: str = Depends(get_current_user),
):
    """Get production data from PBI Operations model (all_prod table).

    The XSPOC NodeID (e.g. 'WEU 216') maps to PBI via:
      Pumper Data[Short Name] -> Pumper Data[Well Name] -> all_prod[M_PROPNUM]
    """
    token = _get_pbi_token()
    from pbi_helpers import execute_dax

    # Look up PBI well name from XSPOC NodeID via Pumper Data
    lookup_dax = f"""
    EVALUATE
    SELECTCOLUMNS(
        FILTER('Pumper Data', 'Pumper Data'[Short Name] = "{well}"),
        "wn", 'Pumper Data'[Well Name]
    )
    """
    try:
        lookup = execute_dax(token, lookup_dax)
    except Exception as e:
        log.error("PBI lookup error: %s", e)
        raise HTTPException(502, f"PBI lookup failed: {e}")

    pbi_name = None
    if lookup:
        pbi_name = lookup[0].get("wn") or lookup[0].get("[wn]")
    if not pbi_name:
        # Fallback: try the well name as-is (might work for non-abbreviated names)
        pbi_name = well

    log.info("PBI lookup: %s -> %s", well, pbi_name)

    # Build DAX query
    cols = PROD_COLUMNS
    if params:
        requested = set(p.strip() for p in params.split(","))
        cols = {k: v for k, v in cols.items() if k in requested}

    col_exprs = ", ".join(f'"{k}", \'all_prod\'[{col}]' for k, (col, _) in cols.items())
    cutoff = datetime.now() - timedelta(days=days)
    dax = f"""
    EVALUATE
    SELECTCOLUMNS(
        FILTER(
            'all_prod',
            'all_prod'[M_PROPNUM] = "{pbi_name}"
                && 'all_prod'[DP_D_DATE] >= DATE({cutoff.year}, {cutoff.month}, {cutoff.day})
        ),
        "dt", 'all_prod'[DP_D_DATE],
        {col_exprs}
    )
    ORDER BY [dt]
    """

    try:
        rows = execute_dax(token, dax)
    except Exception as e:
        log.error("PBI query error: %s", e)
        raise HTTPException(502, f"PBI query failed: {e}")

    # Convert to chart format
    param_data = {}
    for row in rows:
        dt_str = row.get("dt") or row.get("[dt]", "")
        if not dt_str:
            continue
        epoch_ms = int(datetime.fromisoformat(dt_str.replace("Z", "")).timestamp() * 1000)
        for pname, (col, unit) in cols.items():
            val = row.get(pname) or row.get(f"[{pname}]", 0)
            if val is None:
                val = 0
            entry = param_data.setdefault(pname, {"unit": unit, "data": []})
            entry["data"].append([epoch_ms, round(float(val), 2)])

    return {
        "well": well,
        "source": "pbi",
        "days": days,
        "params": param_data,
    }


# ---------------------------------------------------------------------------
# Wellbore Diagram (WBD) endpoints — Ops_Reporting dbo.wbd_* tables
# ---------------------------------------------------------------------------
import re as _re

_ABBREV = {
    'WEST EUMONT UNIT': 'WEU', 'WEST EUMONT': 'WEU',
    'EAST EUMONT UNIT': 'EEU', 'EAST EUMONT': 'EEU',
    'NORTH EUMONT UNIT': 'NEU', 'SOUTH EUMONT UNIT': 'SEU',
    'SOUTH JAL UNIT': 'SJU', 'NORTH JAL UNIT': 'NJU',
    'WHITE CITY UNIT': 'WCU',
}
def _norm(s):
    if not s: return ''
    s = str(s).upper().strip()
    for full, short in _ABBREV.items():
        s = s.replace(full, short)
    s = s.replace('#', ' ')
    s = _re.sub(r'\b0+(\d)', r'\1', s)
    s = _re.sub(r'\s+', ' ', s)
    s = _re.sub(r'[^A-Z0-9 ]', '', s)
    return s.strip()


def _master_short_to_well_id(cur):
    """Build (short -> wbd_wells.well_id) map by name-first matching, API fallback."""
    cur.execute("""
        SELECT [Short Name], [Well Name], [API10 (String)]
        FROM dbo.Pumper_Data_Calcs
        WHERE [Short Name] IS NOT NULL AND LTRIM(RTRIM([Short Name])) <> ''
    """)
    short_by_norm = {}
    short_by_api = {}
    for sh, wn, api in cur.fetchall():
        sh = (sh or '').strip()
        if not sh: continue
        n = _norm(wn)
        if n: short_by_norm[n] = sh
        a = str(api or '').strip()
        if len(a) == 10: short_by_api[a] = sh

    cur.execute("SELECT well_id, well_name, api FROM dbo.wbd_wells")
    short_to_wid = {}
    for wid, wname, api in cur.fetchall():
        n = _norm(wname)
        a = str(api or '').replace('-', '')[:10]
        sh = short_by_norm.get(n) or (short_by_api.get(a) if len(a) == 10 else None)
        if sh and sh not in short_to_wid:
            short_to_wid[sh] = wid
    return short_to_wid


@app.get("/wbd/wells")
def wbd_wells():
    # Unauthenticated read; CORS limits origins. WBD data isn't sensitive.
    """Return all wells (short name + summary fields) for picker autocomplete.

    Includes a `has_wbd` flag so client can mark wells with diagram data.
    """
    conn = get_ops_conn()
    cur = conn.cursor()
    short_to_wid = _master_short_to_well_id(cur)

    # Find which wbd_wells have any extracted version
    cur.execute("SELECT DISTINCT well_id FROM dbo.wbd_versions WHERE data_extracted = 1 AND well_id IS NOT NULL")
    wids_with_data = {r[0] for r in cur.fetchall()}

    cur.execute("""
        SELECT [Short Name], [Well Name], [API10 (String)], Company, Status, Type,
               [Pumper Name], Engineer, Stop, Unit, _LeaseType,
               [WBD Link], WBD_PDFLINK, WellFileLink, Location
        FROM dbo.Pumper_Data_Calcs
        WHERE [Short Name] IS NOT NULL AND LTRIM(RTRIM([Short Name])) <> ''
    """)
    out = []
    cols = [d[0] for d in cur.description]
    for r in cur.fetchall():
        rec = dict(zip(cols, r))
        sh = (rec['Short Name'] or '').strip()
        wid = short_to_wid.get(sh)
        out.append({
            "sn": sh,
            "fn": rec['Well Name'],
            "api": str(rec['API10 (String)'] or ''),
            "co": rec['Company'] or '',
            "stat": rec['Status'] or '',
            "ty": rec['Type'] or '',
            "loc": rec['Location'] or '',
            "has_wbd": (wid in wids_with_data) if wid else False,
        })
    conn.close()
    return {"wells": out, "count": len(out)}


@app.get("/wbd/whoami")
async def whoami_pre(authorization: Optional[str] = Header(None)):
    """Return the caller's email + role. Used by the UI to decide whether to show edit buttons.
    Defined here (before /wbd/{short}) so it doesn't get shadowed by the path-param route."""
    try:
        user = await get_current_user(authorization)
    except HTTPException:
        return {"email": None, "role": None, "can_edit": False}
    role = _fetch_accounts().get(user.lower(), "")
    return {"email": user, "role": role, "can_edit": role in EDITOR_ROLES}


@app.get("/wbd/{short}", response_class=ORJSONResponse)
def wbd_well(short: str):
    """Return full WBD detail for a single well (all extracted versions + components)."""
    conn = get_ops_conn()
    cur = conn.cursor()

    # Pumper master
    cur.execute("""
        SELECT [Short Name], [Well Name], [API10 (String)], Company, Status, Type,
               [Pumper Name], Engineer, Stop, Unit, _LeaseType,
               [WBD Link], WBD_PDFLINK, WellFileLink, Location
        FROM dbo.Pumper_Data_Calcs WHERE [Short Name] = ?
    """, short)
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, f"Well '{short}' not found")
    cols = [d[0] for d in cur.description]
    rec = dict(zip(cols, row))

    short_to_wid = _master_short_to_well_id(cur)
    wid = short_to_wid.get(short)

    well = {
        "fn": rec['Well Name'], "sn": short,
        "api": str(rec['API10 (String)'] or ''),
        "loc": rec['Location'] or '', "co": rec['Company'] or '',
        "stat": rec['Status'] or '', "ty": rec['Type'] or '',
        "pmpr": rec['Pumper Name'] or '', "engr": rec['Engineer'] or '',
        "stop": rec['Stop'] or '', "unit": rec['Unit'] or '',
        "lt": rec['_LeaseType'] or '',
        "wbd_link": rec['WBD Link'] or '', "wbd_pdf": rec['WBD_PDFLINK'] or '',
        "well_link": rec['WellFileLink'] or '',
        "v": [],            # schematic versions (casings/perfs/plugs/...)
        "events": [],       # workover events (rod/tubing/pump from Job Detail)
    }

    # Fetch workover events (independent of whether well_id matches in wbd_wells)
    cur.execute("""
        SELECT we.event_date, we.start_date, we.afe, we.failure_cause, we.secondary_cause,
               we.failed_component, we.component_detail, we.failure_analysis,
               we.pumping_unit, we.spm, we.pump_size, we.pump_type, we.pump_new,
               we.corrosion, we.pull_tubing, we.cleanout, we.acid, we.enduralloy, we.tbg_coating,
               we.tbg_count_278, we.tbg_count_278_coated, we.tbg_count_238, we.tbg_count_238_coated,
               we.tbg_replaced_278, we.tbg_replaced_278_coated, we.tbg_replaced_238, we.tbg_replaced_238_coated,
               we.sn_depth, we.tac_depth,
               we.rods_1, we.rods_78, we.rods_34, we.rods_58, we.rods_fg, we.rods_sinker,
               we.rods_replaced_1, we.rods_replaced_78, we.rods_replaced_34, we.rods_replaced_58, we.rods_replaced_fg, we.rods_replaced_sinker,
               we.rod_couplers_new, we.rod_type, we.summary, we.job_xkey,
               CASE WHEN jd.[Report Path] IS NOT NULL AND LEN(jd.[Report Path]) > 0 THEN 1 ELSE 0 END AS has_report
        FROM dbo.wbd_workover_events we
        LEFT JOIN dbo.[Job Detail] jd ON jd.xKey = we.job_xkey
        WHERE we.sn = ?
        ORDER BY we.event_date DESC
    """, short)
    ev_cols = [d[0] for d in cur.description]
    for r in cur.fetchall():
        e = dict(zip(ev_cols, r))
        # Normalize dates / nulls / boolean
        e["event_date"] = str(e["event_date"])[:10] if e.get("event_date") else None
        e["start_date"] = str(e["start_date"])[:10] if e.get("start_date") else None
        e["has_report"] = bool(e.get("has_report"))
        well["events"].append(e)

    if wid is None:
        conn.close()
        return well

    # Load versions and detail
    cur.execute("""
        SELECT wbd_id, wbd_date, status, td, pbtd
        FROM dbo.wbd_versions
        WHERE well_id = ? AND data_extracted = 1
    """, wid)
    versions = []
    for v in cur.fetchall():
        versions.append({"id": v[0], "dt": str(v[1])[:10] if v[1] else None,
                         "st": v[2] or '', "td": str(v[3]) if v[3] else None,
                         "pb": str(v[4]) if v[4] else None})
    if not versions:
        conn.close()
        return well

    wbd_ids = [v["id"] for v in versions]
    placeholders = ",".join("?" * len(wbd_ids))

    def fetch(sql):
        cur.execute(sql, *wbd_ids)
        out = {}
        cs = [d[0] for d in cur.description]
        for r in cur.fetchall():
            rec = dict(zip(cs, r))
            out.setdefault(rec['wbd_id'], []).append(rec)
        return out

    csg = fetch(f"SELECT * FROM dbo.wbd_casing WHERE wbd_id IN ({placeholders})")
    pf  = fetch(f"SELECT * FROM dbo.wbd_perforations WHERE wbd_id IN ({placeholders})")
    pl  = fetch(f"SELECT * FROM dbo.wbd_plugs WHERE wbd_id IN ({placeholders})")
    ft  = fetch(f"SELECT * FROM dbo.wbd_formation_tops WHERE wbd_id IN ({placeholders})")
    tb  = fetch(f"SELECT * FROM dbo.wbd_tubing WHERE wbd_id IN ({placeholders})")
    rd  = fetch(f"SELECT * FROM dbo.wbd_rods WHERE wbd_id IN ({placeholders})")
    conn.close()

    for v in versions:
        wid2 = v["id"]
        v["csg"] = [{"ty": c.get("casing_type"), "sz": c.get("size"), "wt": c.get("weight"),
                     "gr": c.get("grade"),
                     "dp": str(c["set_depth"]) if c.get("set_depth") else None,
                     "hs": c.get("hole_size"),
                     "cs": str(c["cement_sacks"]) if c.get("cement_sacks") else None,
                     "tc": c.get("toc")} for c in csg.get(wid2, [])]
        v["pf"]  = [{"td": p.get("top_depth"), "bd": p.get("bottom_depth"),
                     "spf": p.get("spf"), "zn": p.get("zone"),
                     "dt": str(p["perf_date"])[:10] if p.get("perf_date") else None,
                     "sq": bool(p["is_squeezed"]) if p.get("is_squeezed") is not None else None}
                    for p in pf.get(wid2, [])]
        v["pl"]  = [{"ty": p.get("plug_type"), "dp": p.get("depth"),
                     "tc": p.get("toc"), "cs": p.get("cement_sacks")} for p in pl.get(wid2, [])]
        v["ft"]  = [{"fm": f.get("formation_name"), "dp": f.get("depth")} for f in ft.get(wid2, [])]
        v["tb"]  = [{"desc": t.get("description"), "dp": t.get("depth"),
                     "jts": t.get("joints")} for t in tb.get(wid2, [])]
        v["rd"]  = [{"desc": r.get("description"), "cnt": r.get("count"),
                     "len": r.get("length"), "dp": r.get("depth")} for r in rd.get(wid2, [])]

    versions.sort(key=lambda x: (0 if x["st"] == "CURRENT" else 1,
                                  -int((x["dt"] or "0000-00-00")[:4]) if x.get("dt") else 0))
    well["v"] = versions
    return well


# ---------------------------------------------------------------------------
# WBD Edit / Add — engineer-only write endpoints
# ---------------------------------------------------------------------------
import urllib.request as _urlreq
import csv as _csv
import io as _io

ACCOUNTS_CSV_URL = "https://docs.google.com/spreadsheets/d/146yuHYjs3RF3wtCK3Qj9NXH94DrStwhieTGqPeLQRfA/gviz/tq?tqx=out:csv&sheet=Accounts"
EDITOR_ROLES = {"engineer", "admin", "tech"}

_accounts_cache = {"data": {}, "ts": 0}
def _fetch_accounts():
    """Return {email_lower: role_lower}. Cached for 5 min."""
    import time as _t
    if _t.time() - _accounts_cache["ts"] < 300 and _accounts_cache["data"]:
        return _accounts_cache["data"]
    try:
        with _urlreq.urlopen(ACCOUNTS_CSV_URL, timeout=10) as r:
            text = r.read().decode("utf-8", errors="replace")
        rows = list(_csv.reader(_io.StringIO(text)))
        out = {}
        if rows:
            headers = [h.lower().strip() for h in rows[0]]
            try:
                ei = headers.index("email")
                ri = headers.index("role")
            except ValueError:
                return {}
            for r in rows[1:]:
                if len(r) <= max(ei, ri): continue
                em = (r[ei] or "").strip().lower()
                ro = (r[ri] or "").strip().lower()
                if em and em != "x":
                    out[em] = ro
        _accounts_cache["data"] = out
        _accounts_cache["ts"] = _t.time()
        return out
    except Exception as e:
        log.warning("Could not fetch Accounts CSV: %s", e)
        return _accounts_cache["data"] or {}


async def get_editor(authorization: Optional[str] = Header(None)) -> str:
    """Verify caller and require Engineer/admin role."""
    user = await get_current_user(authorization)
    accounts = _fetch_accounts()
    role = accounts.get(user.lower(), "")
    if role not in EDITOR_ROLES:
        raise HTTPException(403, f"User {user} role '{role}' not permitted to edit WBDs")
    return user


# ---------------------------------------------------------------------------
# Editable child-table schemas
# ---------------------------------------------------------------------------
from pydantic import BaseModel
from typing import List, Any

class CasingRow(BaseModel):
    casing_type: Optional[str] = None
    size: Optional[str] = None
    weight: Optional[str] = None
    grade: Optional[str] = None
    set_depth: Optional[str] = None
    hole_size: Optional[str] = None
    cement_sacks: Optional[str] = None
    toc: Optional[str] = None
    circ_to_surface: Optional[str] = None
    notes: Optional[str] = None

class PerfRow(BaseModel):
    top_depth: Optional[str] = None
    bottom_depth: Optional[str] = None
    spf: Optional[str] = None
    zone: Optional[str] = None
    perf_date: Optional[str] = None
    is_squeezed: Optional[bool] = None
    squeeze_date: Optional[str] = None
    notes: Optional[str] = None

class PlugRow(BaseModel):
    plug_type: Optional[str] = None
    depth: Optional[str] = None
    toc: Optional[str] = None
    cement_sacks: Optional[str] = None
    notes: Optional[str] = None

class FormationTopRow(BaseModel):
    formation_name: Optional[str] = None
    depth: Optional[str] = None

class TubingRow(BaseModel):
    description: Optional[str] = None
    depth: Optional[str] = None
    joints: Optional[str] = None
    notes: Optional[str] = None

class RodRow(BaseModel):
    description: Optional[str] = None
    count: Optional[str] = None
    length: Optional[str] = None
    depth: Optional[str] = None
    notes: Optional[str] = None

class VersionMeta(BaseModel):
    wbd_date: Optional[str] = None
    status: Optional[str] = None
    td: Optional[str] = None
    pbtd: Optional[str] = None

class WbdEditPayload(BaseModel):
    meta: VersionMeta
    casing: List[CasingRow] = []
    perforations: List[PerfRow] = []
    plugs: List[PlugRow] = []
    formation_tops: List[FormationTopRow] = []
    tubing: List[TubingRow] = []
    rods: List[RodRow] = []

class NewVersionRequest(BaseModel):
    short: str
    clone_from_wbd_id: Optional[int] = None  # if set, copy child rows from this version
    meta: VersionMeta


def _next_id(cur, table, pk):
    cur.execute(f"SELECT ISNULL(MAX({pk}), 0) + 1 FROM dbo.{table}")
    return cur.fetchone()[0]


@app.put("/wbd/version/{wbd_id}")
def edit_wbd_version(wbd_id: int, payload: WbdEditPayload, user: str = Depends(get_editor)):
    """Update a WBD version (metadata + replace all child rows)."""
    conn = get_ops_conn()
    cur = conn.cursor()
    # Verify version exists
    cur.execute("SELECT well_id FROM dbo.wbd_versions WHERE wbd_id = ?", wbd_id)
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, f"WBD version {wbd_id} not found")
    well_id = row[0]

    # Update metadata
    m = payload.meta
    cur.execute(
        "UPDATE dbo.wbd_versions SET wbd_date=?, status=?, td=?, pbtd=?, data_extracted=1 WHERE wbd_id=?",
        m.wbd_date, m.status, m.td, m.pbtd, wbd_id
    )

    # Replace child rows: delete + reinsert
    def replace(table, pk, rows, fields):
        cur.execute(f"DELETE FROM dbo.{table} WHERE wbd_id = ?", wbd_id)
        if not rows: return
        nid = _next_id(cur, table, pk)
        cols = ",".join([pk, "well_id", "wbd_id"] + fields)
        placeholders = ",".join(["?"] * (3 + len(fields)))
        sql = f"INSERT INTO dbo.{table} ({cols}) VALUES ({placeholders})"
        for r in rows:
            d = r.dict()
            params = [nid, well_id, wbd_id] + [d.get(f) for f in fields]
            cur.execute(sql, *params)
            nid += 1

    replace("wbd_casing",         "casing_id", payload.casing,
            ["casing_type","size","weight","grade","set_depth","hole_size","cement_sacks","toc","circ_to_surface","notes"])
    replace("wbd_perforations",   "perf_id",   payload.perforations,
            ["top_depth","bottom_depth","spf","zone","perf_date","is_squeezed","squeeze_date","notes"])
    replace("wbd_plugs",          "plug_id",   payload.plugs,
            ["plug_type","depth","toc","cement_sacks","notes"])
    replace("wbd_formation_tops", "top_id",    payload.formation_tops,
            ["formation_name","depth"])
    replace("wbd_tubing",         "tubing_id", payload.tubing,
            ["description","depth","joints","notes"])
    replace("wbd_rods",           "rod_id",    payload.rods,
            ["description","count","length","depth","notes"])

    conn.commit()
    conn.close()
    log.info("WBD %d edited by %s", wbd_id, user)
    return {"ok": True, "wbd_id": wbd_id, "edited_by": user}


@app.post("/wbd/version")
def create_wbd_version(req: NewVersionRequest, user: str = Depends(get_editor)):
    """Create a new WBD version. If clone_from_wbd_id is set, copy child rows."""
    conn = get_ops_conn()
    cur = conn.cursor()

    # Resolve well_id from short name
    short_to_wid = _master_short_to_well_id(cur)
    well_id = short_to_wid.get(req.short)
    if well_id is None:
        # Could create a wbd_wells row, but for now require it to exist
        conn.close()
        raise HTTPException(404, f"Well '{req.short}' not found in wbd_wells; add it manually first")

    new_id = _next_id(cur, "wbd_versions", "wbd_id")
    m = req.meta
    cur.execute(
        "INSERT INTO dbo.wbd_versions (wbd_id, well_id, wbd_date, status, source_file, file_format, sheet_name, td, pbtd, data_extracted) "
        "VALUES (?, ?, ?, ?, NULL, 'manual', NULL, ?, ?, 1)",
        new_id, well_id, m.wbd_date, m.status, m.td, m.pbtd
    )

    # Optionally clone child tables
    if req.clone_from_wbd_id:
        for table, pk, fields in [
            ("wbd_casing",         "casing_id", ["casing_type","size","weight","grade","set_depth","hole_size","cement_sacks","toc","circ_to_surface","notes"]),
            ("wbd_perforations",   "perf_id",   ["top_depth","bottom_depth","spf","zone","perf_date","is_squeezed","squeeze_date","notes"]),
            ("wbd_plugs",          "plug_id",   ["plug_type","depth","toc","cement_sacks","notes"]),
            ("wbd_formation_tops", "top_id",    ["formation_name","depth"]),
            ("wbd_tubing",         "tubing_id", ["description","depth","joints","notes"]),
            ("wbd_rods",           "rod_id",    ["description","count","length","depth","notes"]),
        ]:
            cur.execute(f"SELECT {','.join(fields)} FROM dbo.{table} WHERE wbd_id = ?", req.clone_from_wbd_id)
            src_rows = cur.fetchall()
            if not src_rows: continue
            nid = _next_id(cur, table, pk)
            cols = ",".join([pk, "well_id", "wbd_id"] + fields)
            placeholders = ",".join(["?"] * (3 + len(fields)))
            for r in src_rows:
                params = [nid, well_id, new_id] + list(r)
                cur.execute(f"INSERT INTO dbo.{table} ({cols}) VALUES ({placeholders})", *params)
                nid += 1

    conn.commit()
    conn.close()
    log.info("WBD version %d created (clone_from=%s) by %s", new_id, req.clone_from_wbd_id, user)
    return {"ok": True, "wbd_id": new_id, "created_by": user}


@app.post("/wbd/parse")
async def parse_wbd_file(
    file: UploadFile = File(...),
    short: Optional[str] = Form(None),
    user: str = Depends(get_editor),
):
    """Accept an uploaded WBD file (.xlsx/.xls/.pdf/.pptx/.ppt), run the extraction
    parser, and return the parsed structure for preview before save.

    Does NOT write to SQL — the client previews/edits in the editor and saves via
    POST /wbd/version + PUT /wbd/version/{id}.
    """
    if file is None:
        raise HTTPException(400, "no file uploaded")
    import tempfile, os as _os, sys as _sys, importlib

    # Save upload to temp
    suffix = _os.path.splitext(file.filename or '')[1].lower() or '.xlsx'
    fd, tmp_path = tempfile.mkstemp(suffix=suffix)
    try:
        with _os.fdopen(fd, 'wb') as f:
            content = await file.read()
            f.write(content)

        # Import / reload extract_wbd module (located alongside api_server.py)
        ext_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'extract_wbd.py')
        if not _os.path.exists(ext_path):
            raise HTTPException(500, f"extract_wbd.py not found at {ext_path}")

        spec = importlib.util.spec_from_file_location('extract_wbd', ext_path)
        ext = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(ext)

        # Reset module globals (parsers append to these)
        ext.wells = {}
        ext.wbd_versions = []
        ext.casing_rows = []
        ext.perf_rows = []
        ext.plug_rows = []
        ext.fish_rows = []
        ext.formation_top_rows = []
        ext.tubing_rows = []
        ext.rod_rows = []

        well_label = (short or 'UPLOADED').strip().upper()
        try:
            from pathlib import Path as _Path
            ext.process_wbd_file(_Path(tmp_path), well_label, 'upload')
        except Exception as e:
            raise HTTPException(400, f"Parser failed: {e}")

        # The parser may emit 0+ wbd_versions. Pick the most recent (or only) one.
        if not ext.wbd_versions:
            raise HTTPException(400, "Parser found no WBD data in this file. Wrong template, blank sheets, or unsupported format.")
        # Sort by date desc (None last) and pick first
        def _vk(v):
            d = v.get('wbd_date') or ''
            return (1 if not d else 0, str(d))
        ext.wbd_versions.sort(key=_vk)
        v = ext.wbd_versions[-1]   # most recent
        wbd_id = v.get('wbd_id')

        def filt(rows):
            return [r for r in rows if r.get('wbd_id') == wbd_id]

        def serialize(rows, drop=()):
            out = []
            for r in rows:
                d = {}
                for k, val in r.items():
                    if k in drop or k in ('well_id', 'wbd_id'): continue
                    if val is None: d[k] = None
                    elif isinstance(val, (str, int, float, bool)): d[k] = val
                    else: d[k] = str(val)
                out.append(d)
            return out

        result = {
            "ok": True,
            "filename": file.filename,
            "well_label": well_label,
            "version_count_in_file": len(ext.wbd_versions),
            "meta": {
                "wbd_date": v.get('wbd_date'),
                "status": v.get('status') or 'CURRENT',
                "td": str(v['td']) if v.get('td') is not None else None,
                "pbtd": str(v['pbtd']) if v.get('pbtd') is not None else None,
            },
            "casing":         serialize(filt(ext.casing_rows)),
            "perforations":   serialize(filt(ext.perf_rows)),
            "plugs":          serialize(filt(ext.plug_rows)),
            "formation_tops": serialize(filt(ext.formation_top_rows)),
            "tubing":         serialize(filt(ext.tubing_rows)),
            "rods":           serialize(filt(ext.rod_rows)),
        }
        log.info("WBD parse by %s: %s -> %d csg / %d pf / %d ft / %d pl / %d tb / %d rd",
                 user, file.filename,
                 len(result["casing"]), len(result["perforations"]), len(result["formation_tops"]),
                 len(result["plugs"]), len(result["tubing"]), len(result["rods"]))
        return result
    finally:
        try: _os.unlink(tmp_path)
        except Exception: pass


@app.get("/wbd/{short}/export")
def export_wbd_xlsx(short: str, user: str = Depends(get_current_user)):
    """Export a well's full WBD data as a multi-sheet .xlsx workbook.

    Sheets: Summary, Schematic Versions, Casings, Perforations, Plugs, Formation Tops,
    Tubing, Rods, Workover Events.
    """
    import io as _io
    from fastapi.responses import StreamingResponse
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed on server")

    conn = get_ops_conn()
    cur = conn.cursor()

    # Pumper master
    cur.execute("""
        SELECT [Short Name], [Well Name], [API10 (String)], Company, Status, Type,
               [Pumper Name], Engineer, Stop, Unit, _LeaseType, Location
        FROM dbo.Pumper_Data_Calcs WHERE [Short Name] = ?
    """, short)
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, f"Well '{short}' not found")
    cols = [d[0] for d in cur.description]
    rec = dict(zip(cols, row))

    short_to_wid = _master_short_to_well_id(cur)
    wid = short_to_wid.get(short)

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1C4A1C")
    title_font = Font(bold=True, size=12, color="1C4A1C")

    def add_sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="left")
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        # Auto-size
        for col in ws.columns:
            mx = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(mx + 2, 10), 60)
        return ws

    # Summary sheet
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = f"Well Info Export — {rec['Short Name']}"
    summary["A1"].font = title_font
    summary.merge_cells("A1:D1")
    summary_rows = [
        ("Well Name",  rec["Well Name"]),
        ("Short Name", rec["Short Name"]),
        ("API",        rec["API10 (String)"]),
        ("Company",    rec["Company"]),
        ("Status",     rec["Status"]),
        ("Type",       rec["Type"]),
        ("Pumper",     rec["Pumper Name"]),
        ("Engineer",   rec["Engineer"]),
        ("Stop",       rec["Stop"]),
        ("Unit",       rec["Unit"]),
        ("Lease Type", rec["_LeaseType"]),
        ("Location",   rec["Location"]),
        ("",           ""),
        ("Exported",   datetime.now().strftime("%Y-%m-%d %H:%M:%S") + f" by {user}"),
    ]
    for ri, (k, v) in enumerate(summary_rows, 3):
        summary.cell(row=ri, column=1, value=k).font = Font(bold=True)
        summary.cell(row=ri, column=2, value=str(v) if v is not None else "")
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 50

    # Schematic versions
    if wid is not None:
        cur.execute("""
            SELECT wbd_id, wbd_date, status, td, pbtd, source_file, file_format, data_extracted
            FROM dbo.wbd_versions WHERE well_id = ? ORDER BY wbd_date DESC
        """, wid)
        v_rows = [tuple(r) for r in cur.fetchall()]
        add_sheet("Schematic Versions",
                  ["wbd_id","wbd_date","status","td","pbtd","source_file","file_format","data_extracted"],
                  v_rows)

        wbd_ids = [r[0] for r in v_rows if r[7]]   # only extracted versions for child detail
        if wbd_ids:
            placeholders = ",".join("?" * len(wbd_ids))

            def fetch_all(table, fields):
                cur.execute(f"SELECT wbd_id, {','.join(fields)} FROM dbo.{table} WHERE wbd_id IN ({placeholders}) ORDER BY wbd_id", *wbd_ids)
                return [tuple(r) for r in cur.fetchall()]

            add_sheet("Casings",
                      ["wbd_id","casing_type","size","weight","grade","set_depth","hole_size","cement_sacks","toc","circ_to_surface","notes"],
                      fetch_all("wbd_casing", ["casing_type","size","weight","grade","set_depth","hole_size","cement_sacks","toc","circ_to_surface","notes"]))
            add_sheet("Perforations",
                      ["wbd_id","top_depth","bottom_depth","spf","zone","perf_date","is_squeezed","squeeze_date","notes"],
                      fetch_all("wbd_perforations", ["top_depth","bottom_depth","spf","zone","perf_date","is_squeezed","squeeze_date","notes"]))
            add_sheet("Plugs",
                      ["wbd_id","plug_type","depth","toc","cement_sacks","notes"],
                      fetch_all("wbd_plugs", ["plug_type","depth","toc","cement_sacks","notes"]))
            add_sheet("Formation Tops",
                      ["wbd_id","formation_name","depth"],
                      fetch_all("wbd_formation_tops", ["formation_name","depth"]))
            add_sheet("Tubing",
                      ["wbd_id","description","depth","joints","notes"],
                      fetch_all("wbd_tubing", ["description","depth","joints","notes"]))
            add_sheet("Rods",
                      ["wbd_id","description","count","length","depth","notes"],
                      fetch_all("wbd_rods", ["description","count","length","depth","notes"]))

    # Workover events
    cur.execute("""
        SELECT event_date, start_date, afe, failure_cause, secondary_cause,
               failed_component, component_detail, failure_analysis,
               pumping_unit, spm, pump_size, pump_type, pump_new,
               sn_depth, tac_depth,
               rods_1, rods_78, rods_34, rods_58, rods_fg, rods_sinker, rod_type, rod_couplers_new,
               tbg_count_278, tbg_count_278_coated, tbg_count_238, tbg_count_238_coated, tbg_coating,
               summary
        FROM dbo.wbd_workover_events WHERE sn = ? ORDER BY event_date DESC
    """, short)
    e_cols = [d[0] for d in cur.description]
    e_rows = [tuple(r) for r in cur.fetchall()]
    add_sheet("Workover Events", e_cols, e_rows)

    conn.close()

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = "".join(c if c.isalnum() else "_" for c in short).strip("_")
    fname = f"WBD_{safe}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


# Map view-friendly field names → actual [Job Detail] columns
WORKOVER_FIELD_MAP = {
    "well_name":              '[Well Name]',
    "failure_date":           '[Failure Date]',
    "start_date":             '[Start]',
    "end_date":               '[End]',
    "afe":                    '[Job/AFE]',
    "summary":                '[Post Job Summary]',
    "failure_cause":          '[Failure Cause]',
    "secondary_cause":        '[Secondary Cause]',
    "failed_component":       '[Failed Component]',
    "component_detail":       '[Component Detail]',
    "failure_analysis":       '[Failure Analysis/Changes]',
    "pumping_unit":           '[Pumping Unit]',
    "spm":                    '[SPM]',
    "rt":                     '[RT]',
    "max_rod_stress":         '[Max Rod Stress]',
    "pro_ject":               '[Pro-Ject?]',
    "review":                 '[Review?]',
    "corrosion":              '[Corrosion?]',
    "pull_tubing":            '[Pull Tubing?]',
    "cleanout":               '[Cleanout?]',
    "acid":                   '[Acid?]',
    "enduralloy":             '[Enduralloy?]',
    "tbg_coating":            '[Tubing Joints_ Coating]',
    "tbg_count_278":          '[Tubing Joints Count 2-7/8"]',
    "tbg_count_278_coated":   '[Tubing Joints Count 2-7/8" Coated]',
    "tbg_count_238":          '[Tubing Joints Count 2-3/8"]',
    "tbg_count_238_coated":   '[Tubing Joints Count 2-3/8" Coated]',
    "tbg_replaced_278":       '[Tubing Joints Replaced 2-7/8"]',
    "tbg_replaced_278_coated":'[Tubing Joints Replaced 2-7/8" Coated]',
    "tbg_replaced_238":       '[Tubing Joints Replaced 2-3/8"]',
    "tbg_replaced_238_coated":'[Tubing Joints Replaced 2-3/8" Coated]',
    "sn_depth":               '[SN Depth]',
    "tac_depth":              '[TAC Depth]',
    "rods_1":                 '[Rods Count 1"]',
    "rods_78":                '[Rods  7/8"]',
    "rods_34":                '[Rods  3/4"]',
    "rods_58":                '[Rods  5/8"]',
    "rods_fg":                '[Rods  FG]',
    "rods_sinker":            '[Rods  SINKER]',
    "rods_replaced_1":        '[Rods Replaced 1"]',
    "rods_replaced_78":       '[Rods  7/8"_1]',
    "rods_replaced_34":       '[Rods  3/4"_2]',
    "rods_replaced_58":       '[Rods  5/8"_3]',
    "rods_replaced_fg":       '[Rods  FG_4]',
    "rods_replaced_sinker":   '[Rods  SINKER_5]',
    "rod_couplers_new":       '[Rods Couplers New (#)]',
    "rod_type":               '[Rods  Type]',
    "pump_size":              '[Pump  Size]',
    "pump_type":              '[Pump  Type]',
    "pump_new":               '[Pump  New?]',
    "xkey":                   '[xKey]',
}


def _compute_xkey(well_name, end_date):
    """Match the dataflow's xKey formula:  UPPER(Well Name) + ' | ' + Date.ToText([End])
    Date format used by Power Query default for en-US is M/d/yyyy."""
    if not well_name or not end_date:
        return None
    try:
        from datetime import datetime as _dt
        if isinstance(end_date, str):
            d = _dt.strptime(end_date[:10], "%Y-%m-%d").date()
        else:
            d = end_date
        date_str = f"{d.month}/{d.day}/{d.year}"
    except Exception:
        date_str = str(end_date)
    return (well_name.strip().upper() + " | " + date_str).upper()


def _next_afe(cur, end_date):
    """Auto-generate AFE for AI/page-driven entries.

    Human AFEs follow YYMM## (6-digit) where ## is a per-(YY,MM) counter.
    AI-generated AFEs prefix with a single 9 -> 9YYMM## (7-digit) so they're
    obviously machine-created. Most jobs should NOT be created this way -
    they should come through the normal workover-report pipeline.
    """
    try:
        from datetime import datetime as _dt
        d = _dt.strptime(end_date[:10], "%Y-%m-%d").date()
    except Exception:
        return None
    yymm = f"{d.year % 100:02d}{d.month:02d}"   # e.g. '2602'
    base_human = int(yymm + "00")               # 260200
    base_ai    = int("9" + yymm + "00")         # 9260200
    cur.execute(
        "SELECT [Job/AFE] FROM dbo.[Job Detail] WHERE "
        "([Job/AFE] BETWEEN ? AND ?) OR ([Job/AFE] BETWEEN ? AND ?)",
        base_human, base_human + 100, base_ai, base_ai + 100
    )
    used_counters = set()
    for r in cur.fetchall():
        if r[0] is None: continue
        v = int(r[0])
        if base_human < v < base_human + 100:
            used_counters.add(v - base_human)
        elif base_ai < v < base_ai + 100:
            used_counters.add(v - base_ai)
    for i in range(1, 100):
        if i not in used_counters:
            return float(base_ai + i)
    return None


def _write_extended(cur, xkey, well_id, ext, source="manual"):
    """Write extended-action rows (perfs/plugs/liners/stim) tagged with given source.
    Idempotent for that source: deletes prior rows of same source for this xkey first."""
    if not ext:
        return
    # Wipe prior rows we own for this xkey
    cur.execute("DELETE FROM dbo.wbd_perforations WHERE add_xkey = ? AND source = ?", xkey, source)
    cur.execute("DELETE FROM dbo.wbd_perforations WHERE squeeze_xkey = ? AND source = ?", xkey, source)
    cur.execute("UPDATE dbo.wbd_perforations SET squeeze_xkey = NULL, squeeze_date = NULL "
                "WHERE squeeze_xkey = ? AND source <> ?", xkey, source)
    cur.execute("DELETE FROM dbo.wbd_plugs WHERE set_xkey = ? AND source = ?", xkey, source)
    cur.execute("DELETE FROM dbo.wbd_plugs WHERE drilled_xkey = ? AND source = ?", xkey, source)
    cur.execute("UPDATE dbo.wbd_plugs SET drilled_xkey = NULL, drilled_date = NULL "
                "WHERE drilled_xkey = ? AND source <> ?", xkey, source)
    cur.execute("DELETE FROM dbo.wbd_casing WHERE set_xkey = ? AND source = ?", xkey, source)
    cur.execute("UPDATE dbo.wbd_casing SET pulled_xkey = NULL, pulled_date = NULL "
                "WHERE pulled_xkey = ? AND source <> ?", xkey, source)
    cur.execute("DELETE FROM dbo.wbd_job_stim WHERE xkey = ? AND source = ?", xkey, source)

    end_date = ext.get("end_date")  # for stamping date columns
    job_type = ext.get("job_type")
    notes    = ext.get("notes")
    cur.execute("DELETE FROM dbo.wbd_job_extended WHERE xkey = ?", xkey)
    if job_type or notes:
        cur.execute("INSERT INTO dbo.wbd_job_extended (xkey, well_id, job_type, source, notes) "
                    "VALUES (?,?,?,?,?)", xkey, well_id, job_type, source, notes)

    def next_id(table, pk):
        cur.execute(f"SELECT ISNULL(MAX({pk}), 0) + 1 FROM dbo.{table}")
        return cur.fetchone()[0]

    for p in ext.get("perfs_added") or []:
        nid = next_id("wbd_perforations", "perf_id")
        cur.execute(
            "INSERT INTO dbo.wbd_perforations (perf_id, well_id, top_depth, bottom_depth, spf, zone, perf_date, add_xkey, source) "
            "VALUES (?,?,?,?,?,?,?,?,?)",
            nid, well_id, p.get("top_depth"), p.get("bottom_depth"), p.get("spf"), p.get("zone"),
            p.get("perf_date") or end_date, xkey, source,
        )
    for p in ext.get("perfs_squeezed") or []:
        # Try to UPDATE an existing perf row for this well that overlaps. Falls back to insert.
        try:
            cur.execute(
                "UPDATE dbo.wbd_perforations SET squeeze_xkey=?, squeeze_date=?, is_squeezed=1 "
                "WHERE well_id=? AND TRY_CAST(top_depth AS INT) <= ? AND TRY_CAST(bottom_depth AS INT) >= ? "
                "AND (squeeze_xkey IS NULL OR squeeze_xkey = ?)",
                xkey, p.get("squeeze_date") or end_date, well_id,
                int(str(p.get("bottom_depth") or "0").split('.')[0]) or 0,
                int(str(p.get("top_depth") or "99999").split('.')[0]) or 99999,
                xkey,
            )
            updated = cur.rowcount
        except Exception:
            updated = 0
        if not updated:
            nid = next_id("wbd_perforations", "perf_id")
            cur.execute(
                "INSERT INTO dbo.wbd_perforations (perf_id, well_id, top_depth, bottom_depth, spf, squeeze_xkey, squeeze_date, is_squeezed, source) "
                "VALUES (?,?,?,?,?,?,?,1,?)",
                nid, well_id, p.get("top_depth"), p.get("bottom_depth"), p.get("spf"),
                xkey, p.get("squeeze_date") or end_date, source,
            )
    for pl in ext.get("plugs_set") or []:
        nid = next_id("wbd_plugs", "plug_id")
        cur.execute(
            "INSERT INTO dbo.wbd_plugs (plug_id, well_id, plug_type, depth, toc, cement_sacks, set_xkey, set_date, source, notes) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            nid, well_id, pl.get("plug_type"), pl.get("depth"), pl.get("toc"), pl.get("cement_sacks"),
            xkey, pl.get("set_date") or end_date, source, pl.get("notes"),
        )
    for pl in ext.get("plugs_drilled") or []:
        try:
            cur.execute(
                "UPDATE dbo.wbd_plugs SET drilled_xkey=?, drilled_date=? "
                "WHERE well_id=? AND plug_type=? AND TRY_CAST(depth AS INT) = ? AND drilled_xkey IS NULL",
                xkey, pl.get("drilled_date") or end_date, well_id, pl.get("plug_type"),
                int(str(pl.get("depth") or "0").split('.')[0]) or 0,
            )
            updated = cur.rowcount
        except Exception:
            updated = 0
        if not updated:
            nid = next_id("wbd_plugs", "plug_id")
            cur.execute(
                "INSERT INTO dbo.wbd_plugs (plug_id, well_id, plug_type, depth, drilled_xkey, drilled_date, source, notes) "
                "VALUES (?,?,?,?,?,?,?,?)",
                nid, well_id, pl.get("plug_type"), pl.get("depth"),
                xkey, pl.get("drilled_date") or end_date, source, pl.get("notes"),
            )
    for l in ext.get("liners_installed") or []:
        nid = next_id("wbd_casing", "casing_id")
        cur.execute(
            "INSERT INTO dbo.wbd_casing (casing_id, well_id, casing_type, size, weight, set_depth, hole_size, cement_sacks, toc, set_xkey, set_date, source, notes) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            nid, well_id, "Liner", l.get("size"), l.get("weight"),
            l.get("set_bottom") or l.get("set_depth"), l.get("hole_size"),
            l.get("cement_sxs") or l.get("cement_sacks"), l.get("toc"),
            xkey, l.get("set_date") or end_date, source, l.get("notes"),
        )
    for l in ext.get("liners_pulled") or []:
        cur.execute(
            "UPDATE dbo.wbd_casing SET pulled_xkey=?, pulled_date=? "
            "WHERE well_id=? AND casing_type='Liner' AND pulled_xkey IS NULL",
            xkey, l.get("pulled_date") or end_date, well_id,
        )
    for s in ext.get("stimulation") or []:
        cur.execute(
            "INSERT INTO dbo.wbd_job_stim (xkey, well_id, stim_type, volume, volume_unit, top_depth, bottom_depth, stim_date, source, notes) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            xkey, well_id, s.get("stim_type"), s.get("volume"), s.get("volume_unit"),
            s.get("top_depth"), s.get("bottom_depth"), s.get("stim_date") or end_date, source, s.get("notes"),
        )


# ---------------------------------------------------------------------------
# Workover-report download — exposes the original XLSX file for a job. The
# stored path in [Job Detail].[Report Path] is in 'dbx:/...' form (Dropbox-
# relative); this endpoint downloads via Dropbox API and streams to the
# browser. Auth-gated like everything else.
#
# Dropbox creds expected at <api_server.py dir>/config.json with a "dropbox"
# block (app_key, app_secret, refresh_token, root_namespace_id).
# ---------------------------------------------------------------------------
_DBX_CFG_API = None
_DBX_TOKEN_API = None

def _load_api_dropbox_cfg():
    global _DBX_CFG_API
    if _DBX_CFG_API is not None:
        return _DBX_CFG_API
    try:
        cfg_path = Path(__file__).parent / "config.json"
        if cfg_path.is_file():
            with open(cfg_path) as f:
                _DBX_CFG_API = (json.load(f).get("dropbox") or {})
        else:
            _DBX_CFG_API = {}
    except Exception as e:
        log.warning("Could not load Dropbox config: %s", e)
        _DBX_CFG_API = {}
    return _DBX_CFG_API


def _dbx_token_api():
    global _DBX_TOKEN_API
    if _DBX_TOKEN_API:
        return _DBX_TOKEN_API
    cfg = _load_api_dropbox_cfg()
    if not cfg.get("refresh_token"):
        return None
    import requests
    resp = requests.post("https://api.dropboxapi.com/oauth2/token", data={
        "grant_type":    "refresh_token",
        "refresh_token": cfg["refresh_token"],
        "client_id":     cfg["app_key"],
        "client_secret": cfg["app_secret"],
    }, timeout=15)
    resp.raise_for_status()
    _DBX_TOKEN_API = resp.json()["access_token"]
    return _DBX_TOKEN_API


@app.get("/wbd/workover/{xkey:path}/report-link")
def get_workover_report_link(xkey: str, user: str = Depends(get_current_user)):
    """Return a short-lived (4-hour) direct Dropbox URL for the workover report.

    Client can then window.open() the URL — no streaming through this server.
    """
    import requests as _r
    conn = get_ops_conn()
    cur = conn.cursor()
    cur.execute("SELECT [Report Path] FROM dbo.[Job Detail] WHERE xKey = ?", xkey)
    r = cur.fetchone()
    conn.close()
    if not r or not r[0]:
        raise HTTPException(404, "no report path on file for this job")
    stored = str(r[0]).strip()
    if not stored.lower().startswith("dbx:"):
        raise HTTPException(404, "stored path is not in dbx: form")
    dbx_path = stored[4:]
    cfg = _load_api_dropbox_cfg()
    tok = _dbx_token_api()
    if not tok:
        raise HTTPException(500, "Dropbox credentials not configured on this server")
    headers = {"Authorization": f"Bearer {tok}", "Content-Type": "application/json"}
    ns = cfg.get("root_namespace_id")
    if ns:
        headers["Dropbox-API-Path-Root"] = json.dumps({".tag": "root", "root": ns})
    try:
        resp = _r.post(
            "https://api.dropboxapi.com/2/files/get_temporary_link",
            headers=headers, json={"path": dbx_path}, timeout=15,
        )
        if resp.status_code >= 400:
            log.warning("Dropbox temp-link %s -> %d %s", dbx_path, resp.status_code, resp.text[:200])
            raise HTTPException(502, f"Dropbox returned {resp.status_code}")
        data = resp.json()
    except _r.RequestException as e:
        raise HTTPException(502, f"Dropbox temp-link failed: {e}")
    fname = os.path.basename(dbx_path) or "workover_report.xlsx"
    return {"link": data.get("link"), "filename": fname}


@app.get("/wbd/workover/{xkey:path}/report")
def download_workover_report(xkey: str, user: str = Depends(get_current_user)):
    """Stream the original workover-report XLSX for a job. Resolves
    [Job Detail].[Report Path] (dbx:/... form) via Dropbox API and pipes the
    file body back to the browser. 404 if the row has no path or it's a
    legacy local-only path the server can't reach.
    """
    from fastapi.responses import StreamingResponse
    import requests as _r
    conn = get_ops_conn()
    cur = conn.cursor()
    cur.execute("SELECT [Report Path] FROM dbo.[Job Detail] WHERE xKey = ?", xkey)
    r = cur.fetchone()
    conn.close()
    if not r or not r[0]:
        raise HTTPException(404, "no report path on file for this job")
    stored = str(r[0]).strip()
    if not stored.lower().startswith("dbx:"):
        raise HTTPException(404, "stored path is not in dbx: form; can't fetch from Dropbox")
    dbx_path = stored[4:]
    cfg = _load_api_dropbox_cfg()
    tok = _dbx_token_api()
    if not tok:
        raise HTTPException(500, "Dropbox credentials not configured on this server")
    headers = {"Authorization": f"Bearer {tok}",
               "Dropbox-API-Arg": json.dumps({"path": dbx_path})}
    ns = cfg.get("root_namespace_id")
    if ns:
        headers["Dropbox-API-Path-Root"] = json.dumps({".tag": "root", "root": ns})
    try:
        dbx_resp = _r.post("https://content.dropboxapi.com/2/files/download",
                           headers=headers, stream=True, timeout=60)
        if dbx_resp.status_code >= 400:
            log.warning("Dropbox download %s -> %d %s", dbx_path, dbx_resp.status_code, dbx_resp.text[:200])
            raise HTTPException(502, f"Dropbox returned {dbx_resp.status_code}")
    except _r.RequestException as e:
        raise HTTPException(502, f"Dropbox download failed: {e}")
    fname = os.path.basename(dbx_path) or "workover_report.xlsx"
    log.info("Streaming workover report %s for %s (xkey=%s)", fname, user, xkey)
    return StreamingResponse(
        dbx_resp.iter_content(chunk_size=64 * 1024),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/wbd/workover/{xkey:path}/extended")
def get_workover_extended(xkey: str, user: str = Depends(get_current_user)):
    """Return all extended rows (perfs/plugs/liners/stim/job_type) for one xKey."""
    conn = get_ops_conn()
    cur = conn.cursor()
    out = {"xkey": xkey}
    def rows(sql, *params):
        cur.execute(sql, *params)
        cs = [d[0] for d in cur.description]
        return [dict(zip(cs, r)) for r in cur.fetchall()]

    out["job_type"] = None
    cur.execute("SELECT job_type, source, notes FROM dbo.wbd_job_extended WHERE xkey = ?", xkey)
    r = cur.fetchone()
    if r:
        out["job_type"] = r[0]; out["job_extended_source"] = r[1]; out["notes"] = r[2]

    out["perfs_added"]      = rows("SELECT perf_id, top_depth, bottom_depth, spf, zone, perf_date, source FROM dbo.wbd_perforations WHERE add_xkey = ?", xkey)
    out["perfs_squeezed"]   = rows("SELECT perf_id, top_depth, bottom_depth, spf, zone, squeeze_date, source FROM dbo.wbd_perforations WHERE squeeze_xkey = ?", xkey)
    out["plugs_set"]        = rows("SELECT plug_id, plug_type, depth, toc, cement_sacks, set_date, source, notes FROM dbo.wbd_plugs WHERE set_xkey = ?", xkey)
    out["plugs_drilled"]    = rows("SELECT plug_id, plug_type, depth, drilled_date, source, notes FROM dbo.wbd_plugs WHERE drilled_xkey = ?", xkey)
    out["liners_installed"] = rows("SELECT casing_id, size, weight, set_depth as set_bottom, hole_size, cement_sacks as cement_sxs, toc, set_date, source, notes FROM dbo.wbd_casing WHERE set_xkey = ? AND casing_type = 'Liner'", xkey)
    out["liners_pulled"]    = rows("SELECT casing_id, size, set_depth as set_bottom, pulled_date, source FROM dbo.wbd_casing WHERE pulled_xkey = ? AND casing_type = 'Liner'", xkey)
    out["stimulation"]      = rows("SELECT stim_id, stim_type, volume, volume_unit, top_depth, bottom_depth, stim_date, source, notes FROM dbo.wbd_job_stim WHERE xkey = ?", xkey)
    conn.close()
    return out


@app.put("/wbd/workover/{xkey:path}/extended")
def edit_workover_extended(xkey: str, payload: dict, user: str = Depends(get_editor)):
    """Manual edit of extended rows. Writes with source='manual' (preserves narrative_parser rows for other workflows is the intent — but we wipe-and-rewrite all manual rows for this xkey on save)."""
    conn = get_ops_conn()
    cur = conn.cursor()
    # Resolve well_id
    cur.execute("SELECT well_id FROM dbo.wbd_job_extended WHERE xkey = ?", xkey)
    r = cur.fetchone()
    well_id = r[0] if r else None
    if well_id is None:
        # Fall back via Job Detail's well_name
        cur.execute(
            "SELECT TOP 1 ww.well_id FROM dbo.[Job Detail] jd "
            "JOIN dbo.Pumper_Data_Calcs pdc ON UPPER(jd.[Well Name]) = UPPER(pdc.[Well Name]) "
            "                                OR UPPER(jd.[Well Name]) = UPPER(pdc.[Short Name]) "
            "JOIN dbo.wbd_wells ww ON UPPER(ww.well_name) = UPPER(pdc.[Well Name]) OR UPPER(ww.well_name) = UPPER(pdc.[Short Name]) "
            "WHERE jd.xKey = ?", xkey
        )
        r = cur.fetchone()
        well_id = r[0] if r else None

    _write_extended(cur, xkey, well_id, payload, source="manual")
    conn.commit()
    conn.close()
    log.info("Extended (manual) updated for %s by %s", xkey, user)
    return {"ok": True, "xkey": xkey}


@app.post("/wbd/workover")
def create_or_upsert_workover(payload: dict, user: str = Depends(get_editor)):
    """Create a new Job Detail row (or upsert if xKey already exists).

    Accepts the same field names as /wbd/workover/{xkey} PUT plus 'short' (well short
    name) for resolving the well. If 'afe' is missing, auto-generates from end_date
    in YYMMDD format (with .N decimal bumps on collision).
    """
    from datetime import datetime as _dt
    if not payload:
        raise HTTPException(400, "empty payload")

    short = payload.pop("short", None)
    well_name = payload.get("well_name")
    end_date = payload.get("end_date")
    if not end_date:
        raise HTTPException(400, "end_date required (YYYY-MM-DD)")

    conn = get_ops_conn()
    cur = conn.cursor()

    # Resolve well_name from short if needed
    if not well_name and short:
        cur.execute("SELECT [Well Name] FROM dbo.Pumper_Data_Calcs WHERE [Short Name] = ?", short)
        r = cur.fetchone()
        if r:
            well_name = r[0]
            payload["well_name"] = well_name
    if not well_name:
        conn.close()
        raise HTTPException(400, "well_name (or 'short') required")

    # Compute xKey + auto-generate AFE if missing
    xkey = _compute_xkey(well_name, end_date)
    payload["xkey"] = xkey
    if not payload.get("afe"):
        afe = _next_afe(cur, end_date)
        if afe is not None:
            payload["afe"] = afe

    # If xKey already exists → update; else insert
    cur.execute("SELECT COUNT(*) FROM dbo.[Job Detail] WHERE xKey = ?", xkey)
    exists = cur.fetchone()[0] > 0

    # Coerce dates
    for k in ("start_date", "end_date", "failure_date"):
        v = payload.get(k)
        if v and isinstance(v, str):
            try: payload[k] = _dt.strptime(v[:10], "%Y-%m-%d").date()
            except Exception: pass
    # Coerce afe to number
    if payload.get("afe") is not None and not isinstance(payload["afe"], (int, float)):
        try: payload["afe"] = float(payload["afe"])
        except Exception: pass

    if exists:
        # Update path
        set_clauses, params = [], []
        for k, v in payload.items():
            col = WORKOVER_FIELD_MAP.get(k)
            if not col or k == "xkey": continue
            set_clauses.append(f"{col} = ?")
            params.append(v)
        if set_clauses:
            sql = f"UPDATE dbo.[Job Detail] SET {', '.join(set_clauses)} WHERE xKey = ?"
            params.append(xkey)
            cur.execute(sql, *params)
        action = "updated"
    else:
        # Insert path
        cols, params = [], []
        for k, v in payload.items():
            col = WORKOVER_FIELD_MAP.get(k)
            if not col: continue
            cols.append(col)
            params.append(v)
        if not cols:
            conn.close()
            raise HTTPException(400, "no recognized fields in payload")
        placeholders = ", ".join(["?"] * len(cols))
        sql = f"INSERT INTO dbo.[Job Detail] ({', '.join(cols)}) VALUES ({placeholders})"
        cur.execute(sql, *params)
        action = "created"

    # Extended action rows (perfs_added/squeezed, plugs_set/drilled, liners_installed/pulled, stimulation, job_type)
    ext_keys = ("perfs_added","perfs_squeezed","plugs_set","plugs_drilled","liners_installed","liners_pulled","stimulation","job_type","notes")
    if any(k in payload for k in ext_keys):
        # Resolve well_id if we don't have it
        cur.execute(
            "SELECT TOP 1 ww.well_id FROM dbo.wbd_wells ww JOIN dbo.Pumper_Data_Calcs pdc "
            " ON UPPER(ww.well_name) = UPPER(pdc.[Well Name]) OR UPPER(ww.well_name) = UPPER(pdc.[Short Name]) "
            "WHERE UPPER(pdc.[Well Name]) = UPPER(?) OR UPPER(pdc.[Short Name]) = UPPER(?)",
            well_name, well_name
        )
        rr = cur.fetchone()
        wid = rr[0] if rr else None
        ext_payload = {k: payload[k] for k in ext_keys if k in payload}
        ext_payload["end_date"] = end_date
        _write_extended(cur, xkey, wid, ext_payload, source="manual")

    conn.commit()
    conn.close()
    log.info("Workover %s by %s: well=%s end=%s afe=%s xkey=%s",
             action, user, well_name, end_date, payload.get("afe"), xkey)
    return {"ok": True, "action": action, "xkey": xkey, "afe": payload.get("afe")}


@app.put("/wbd/workover/{xkey:path}")
def edit_workover(xkey: str, payload: dict, user: str = Depends(get_editor)):
    """Update a Job Detail row matching xKey. Only fields in WORKOVER_FIELD_MAP are accepted."""
    if not payload:
        raise HTTPException(400, "empty payload")
    set_clauses = []
    params = []
    for k, v in payload.items():
        col = WORKOVER_FIELD_MAP.get(k)
        if not col:
            continue
        set_clauses.append(f"{col} = ?")
        params.append(v)
    if not set_clauses:
        raise HTTPException(400, "no editable fields in payload")

    sql = f"UPDATE dbo.[Job Detail] SET {', '.join(set_clauses)} WHERE xKey = ?"
    params.append(xkey)
    conn = get_ops_conn()
    cur = conn.cursor()
    cur.execute(sql, *params)
    if cur.rowcount == 0:
        conn.close()
        raise HTTPException(404, f"Job Detail row with xKey '{xkey}' not found")
    conn.commit()
    conn.close()
    log.info("Workover xKey=%s updated by %s (%d cols)", xkey, user, len(set_clauses))
    return {"ok": True, "xkey": xkey, "updated_fields": list(payload.keys())}


@app.delete("/wbd/version/{wbd_id}")
def soft_delete_wbd_version(wbd_id: int, user: str = Depends(get_editor)):
    """Soft delete: flip data_extracted = 0 so the version disappears from UI but stays in DB."""
    conn = get_ops_conn()
    cur = conn.cursor()
    cur.execute("UPDATE dbo.wbd_versions SET data_extracted = 0 WHERE wbd_id = ?", wbd_id)
    if cur.rowcount == 0:
        conn.close()
        raise HTTPException(404, f"WBD version {wbd_id} not found")
    conn.commit()
    conn.close()
    log.info("WBD %d soft-deleted by %s", wbd_id, user)
    return {"ok": True}


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import uvicorn
    log.info("Starting XSPOC API server on port %d ...", PORT)
    uvicorn.run(app, host="0.0.0.0", port=PORT, log_level="info")
